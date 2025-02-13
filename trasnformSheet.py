import pandas as pd
from tkinter import Tk, Canvas, Entry, Button, filedialog, messagebox
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys



def select_training_file():
    global training_file
    training_file = filedialog.askopenfilename(title="Select Training File", filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if training_file:
        file_name = training_file.split("/")[-3:]
        str_file_name = ""
        for conc in file_name:
            str_file_name += f"/{conc}" 
        display_name = str_file_name if len(str_file_name) <= 60 else str_file_name[-60:]
        btn_training.config(text=f"{display_name}")
        
        
def process_files():
    global training_file

    if not training_file:
        messagebox.showerror("Error", "Select the excel file!")
        return

    try:
        df_training = pd.read_excel(training_file)
        df_email = pd.read_csv('view_base_funcionarios.csv', encoding='windows-1252', delimiter=';')

        df_training['Matricula_temp'] = df_training['Matricula'].astype(str).str.zfill(8)
        df_email['Matrícula do Funcionário'] = df_email['Matrícula do Funcionário'].astype(str).str.zfill(8)

        df_merged = pd.merge(df_training, df_email[['Matrícula do Funcionário', 'Email Gestor', 'E-mail do Funcionário']],
                             left_on='Matricula_temp', right_on='Matrícula do Funcionário', how='left')
        df_merged.drop(columns=['Matricula_temp', 'Matrícula do Funcionário'], inplace=True)

        # Format date columns
        date_columns = ['Data Início', 'Data Fim']
        for col in date_columns:
            if col in df_merged.columns:
                df_merged[col] = pd.to_datetime(df_merged[col], errors='coerce').dt.strftime('%d/%m/%Y')

        output_filename = filename_entry.get().strip()
        if not output_filename:
            messagebox.showerror("Error", "O nome informado não é válido.")
            return

        with pd.ExcelWriter(f"{output_filename}.xlsx", engine='openpyxl') as writer:
            df_merged.to_excel(writer, index=False)
            worksheet = writer.sheets[writer.book.sheetnames[0]]

            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                max_length = 0
                col_letter = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 6
                worksheet.column_dimensions[col_letter].width = adjusted_width

            table = Table(displayName="Tabela1", ref=worksheet.dimensions)
            style = TableStyleInfo(name="TableStyleLight14", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=True)
            table.tableStyleInfo = style
            worksheet.add_table(table)

        messagebox.showinfo("Success", f"File saved as {output_filename}.xlsx!")
        sys.exit()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")


# configs of window
window = Tk()
window.geometry("620x400")
window.configure(bg="#83848B")
window.title("File Processor")

canvas = Canvas(window, bg="#83848B", height=400, width=620, bd=0, highlightthickness=0, relief="ridge")
canvas.place(x=0, y=0)

canvas.create_text(310, 55, anchor="center", text="Formatar Planilha", fill="#FFFFFF", font=("Inter SemiBold", 32))


btn_training = Button(window, text="Procurar arquivo",  command=select_training_file, bg="#4A4F6B", fg="#DADADA", borderwidth=1)
btn_training.place(x=125, y=120, width=368, height=36)

canvas.create_text(125, 185, anchor="nw", text="Nomeie o arquivo", fill="#FFFFFF", font=("Helvetica", 12))

filename_entry = Entry(window, bd=1, bg="#FFFFFF", fg="#000716", justify='center', borderwidth=1, font=("Helvetica", 12))
filename_entry.place(x=125, y=220, width=368, height=36)

btn_process = Button(window, text="Processar", command=process_files, bg="#4A4F6B", fg="#DADADA", borderwidth=1)
btn_process.place(x=209, y=313, width=202, height=38)

window.resizable(False, False)
window.mainloop()
